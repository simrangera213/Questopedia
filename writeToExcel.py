import openpyxl
import os
import string
import actions
import time


def writeToExcel(prediction , query , exec_time):
    wb = openpyxl.load_workbook('score/score.xlsx');
    workbook = wb.active

    if(workbook.max_row > 1):
        data = ( workbook.max_row , query ,  prediction[0] , exec_time , prediction[1] , prediction[2] )
        workbook.append(data)            
    else:
        data1 = ("S.No." , "Question" , "Answer" , "Execution Time" , "Title" , "Paragraph" )
        workbook.append(data1)
        data = (1 , query ,  prediction[0] , exec_time , prediction[1] , prediction[2] )
        workbook.append(data)
        
    wb.save('score/score.xlsx');
    


